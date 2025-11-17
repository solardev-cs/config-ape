import streamlit as st
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
from millify import millify
import locale
import keyboard

import tarifas as tf
import irradiacao as ir
import excel as ex

# AJUSTES FUTUROS PARA WEB APP FUNCIONAL
# 1. Biblioteca xlwings não funciona na web sem Excel instalado no servidor - desenvolver cálculo fora do Excel, em dataframes
# 2. Biblioteca keyboard (ambiente windows) para pressionar tecla não funciona na web - atualizar app programaticamente
# 3. Servidor web precisa aceitar locale pt_BR - outras alternativas alterando software

# PORTUGUÊS DO BRASIL
try:
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
except locale.Error:
    print("Locale 'pt_BR.UTF-8' não disponível, utilizando o padrão do servidor.")

# CONFIGURAÇÃO DA PÁGINA
st.set_page_config(
    page_title="Config APE",
    page_icon="",
    layout="wide",
    initial_sidebar_state="collapsed",
    menu_items={
        'About': "Configurador APE Fockink v1.0"
    }
)

# SUBGRUPOS
subg = ["A2", "A3", "A3a", "A4"]
# MODALIDADES
modal = ["VERDE", "AZUL"]
# ESTRUTURA / TIPO INSTALAÇÃO
estr = ["SOLO FIXO", "SOLO TRACKER", "TELHADO"]
# FATOR DE POTÊNCIA
fatorp = (1, 0.92)
# TIPOS DE PAGAMENTO
tipopag = ["RECURSO PRÓPRIO", "FINANCIAMENTO"]
# TIPOS DE FINANCIAMENTO
tipofinanc = ["SAC", "PRICE"]

# SPINNER CUSTOM
def make_spinner(text):
    with st.spinner(text):
        yield

# OBTÉM PLANILHAS DE LEITURA E ESCRITA
planilha_l = ex.wb_l["APE"]
planilha_e = ex.wb_e["APE"]

# OBTÉM BANCO DE DADOS DE IRRADIAÇÃO
banco_irrad = ir.carrega_bd()

# MODAL CONFIGURAÇÕES
@st.dialog("Configurações")
def config():
    with st.form("form_config", enter_to_submit=False, border=False):
        with st.container():
            coluna1, coluna2 = st.columns(2)
            with coluna1:
                st.markdown("##### Custo Médio (R$/kWp)")
            with coluna2:
                custo = st.number_input("Custo médio", label_visibility="collapsed", value=4000)   
        
        with st.container():
            coluna1, coluna2 = st.columns(2)
            with coluna1:
                st.markdown("##### Perc. Simultaneidade (%)")
            with coluna2:
                perc_simult = st.number_input("Perc simult", label_visibility="collapsed", step=0.1, value=40.0)

        with st.container():
            coluna1, coluna2 = st.columns(2)
            with coluna1:
                st.markdown("##### ICMS (%)")
            with coluna2:
                icms = st.number_input("ICMS", label_visibility="collapsed", step= 0.01, value=18.0) 

        with st.container():
            coluna1, coluna2 = st.columns(2)
            with coluna1:
                st.markdown("##### PIS/COFINS (%)")
            with coluna2:
                pis_cofins = st.number_input("PIS/COFINS", label_visibility="collapsed", step=0.01, value=5.0)

        with st.container():
            coluna1, coluna2 = st.columns(2)
            with coluna1:
                st.markdown("##### ICMS TUSD")
            with coluna2:
                icms_tusd = st.selectbox("ICMS TUSD", (["Sim", "Não"]), label_visibility="collapsed")
        
        with st.container():
            coluna1, coluna2 = st.columns(2)
            with coluna1:
                st.markdown("##### Ganho Tracker (%)")
            with coluna2:
                ganho_tracker = st.number_input("Ganho tracker", label_visibility="collapsed", step=0.1, value=20.0)

        with st.container():
            coluna1, coluna2 = st.columns(2)
            with coluna1:
                st.markdown("##### Sobrecarga Inversor")
            with coluna2:
                sobrecarga = st.number_input("Sobrecarga inversor", label_visibility="collapsed", step=0.01, value=1.4)

        with st.container():
            coluna1, coluna2 = st.columns(2)
            with coluna1:
                st.markdown("##### Pp Módulo (Wp)")
            with coluna2:
                pp_mod = st.number_input("Pp módulo", label_visibility="collapsed", step=1, value=545)

        with st.container():
            coluna1, coluna2 = st.columns(2)
            with coluna1:
                st.markdown("##### Área Módulo (m²)")
            with coluna2:
                area_mod = st.number_input("Área módulo", label_visibility="collapsed", value=2.564)
        
        with st.container():
            coluna1, coluna2 = st.columns(2)
            with coluna1:
                st.markdown("##### Eficiência Módulo (%)")
            with coluna2:
                n_mod = st.number_input("Eficiência módulo", label_visibility="collapsed", step=0.01, value=21.3)

        with st.container():
            coluna1, coluna2 = st.columns(2)
            with coluna1:
                st.markdown("##### Perda Inicial (%)")
            with coluna2:
                perda_0 = st.number_input("Perda inicial", label_visibility="collapsed", step=0.01, value=2.0)
        
        with st.container():
            coluna1, coluna2 = st.columns(2)
            with coluna1:
                st.markdown("##### Perda Anual (%)")
            with coluna2:
                perda_ano = st.number_input("Perda anual", label_visibility="collapsed", step=0.01, value=0.8)
        
        with st.container():
            coluna1, coluna2 = st.columns(2)
            with coluna1:
                st.markdown("##### Ganho Bifacial (%)")
            with coluna2:
                ganho_bif = st.number_input("Ganho bifacial", label_visibility="collapsed", step=0.1)

        with st.container():
            coluna1, coluna2 = st.columns(2)
            with coluna2:
                # Botão para salvar formulário
                submitted6 = st.form_submit_button("Salvar", icon=":material/check:") 

    if submitted6:
        # Cria spinner e inicia
        sp = iter(make_spinner("Salvando"))
        next(sp)

        # Salva variáveis na planilha
        planilha_e["C8"].value = custo
        planilha_e["M17"].value = (perc_simult/100)
        planilha_e["G5"].value = (icms/100)
        planilha_e["G6"].value = (pis_cofins/100)
        planilha_e["G7"].value = icms_tusd
        planilha_e["F85"].value = (ganho_tracker/100)
        planilha_e["F86"].value = sobrecarga
        planilha_e["C81"].value = pp_mod
        planilha_e["C83"].value = area_mod
        planilha_e["C84"].value = n_mod
        planilha_e["C85"].value = (perda_0/100)
        planilha_e["C86"].value = (perda_ano/100)
        planilha_e["C87"].value = (ganho_bif/100)
        
        # Salva planilha e atualiza front-end
        ex.salva_backend()
        
        # Recarrega planilha de leitura, após ter sido atualizada e recalculada
        wb_l = ex.abre_backend_leitura()
        planilha_l = wb_l["APE"]

        # Finaliza spinner
        next(sp, None)

# CABEÇALHO DO APP
with st.container():
    col1,col2,col3 = st.columns([20,10,1])
    with col1:
        st.logo("images/logo.png", size="large", icon_image="images/icone.png")
        st.subheader("CONFIGURADOR APE")
        st.markdown(" <style> div[class^='block-container'] { padding-top: 3rem; } </style> ", unsafe_allow_html=True)
    with col3:
        st.markdown(" <style> div[class^='block-container'] { padding-top: 3rem; text-align: left; } </style> ", unsafe_allow_html=True)
        if st.button("", icon=":material/settings:"):
            config()

# PARTE CENTRAL DO APP
with st.container():

    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(["Dados Iniciais", "Unidade Consumidora", "Usina FV", "Contrato de Energia", "Viabilidade", "Resumo"])
    
    # DADOS INICIAIS
    with tab1: 
        st.write("")

        with st.container():
            coluna1, coluna2 = st.columns(2)
            with coluna1:
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("##### UF")
                with col2:
                    estado = st.selectbox("UF", (banco_irrad['STATE'].unique()), label_visibility="collapsed")

        # Seleciona apenas cidades do respectivo estado
        cidades_filtradas = sorted(banco_irrad[banco_irrad['STATE'] == estado]['NAME'].unique(), key=locale.strxfrm)

        with st.form("form_dadosIniciais", enter_to_submit=False, border=False):
            with st.container():
                col1, col2 = st.columns(2)
                with col1:                           
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:
                            st.markdown("##### Cidade")
                        with coluna2:
                            cidade = st.selectbox("Cidade", (cidades_filtradas), label_visibility="collapsed")                    
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:
                            st.markdown("##### Cliente")
                        with coluna2:
                            cliente = st.text_input("Cliente", label_visibility="collapsed")                     
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:
                            st.markdown("##### Ano de análise")
                        with coluna2:
                            ano_analise = st.number_input("Ano", label_visibility="collapsed", value=0)
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna2:
                            # Botão para salvar formulário
                            submitted = st.form_submit_button("Salvar", icon=":material/check:")            

        # Executa quando botão Salvar é clicado
        if submitted:
            # Cria spinner e inicia
            sp = iter(make_spinner("Salvando"))
            next(sp)
            
            irradiacao = ir.busca_irrad(estado, cidade)

            # Salva variáveis na planilha
            planilha_e["C5"].value = cliente
            planilha_e["C6"].value = estado
            planilha_e["C7"].value = cidade
            planilha_e["C9"].value = ano_analise

            planilha_e["I17"].value = float(irradiacao.JAN.iloc[0]/1000)
            planilha_e["I18"].value = float(irradiacao.FEB.iloc[0]/1000)
            planilha_e["I19"].value = float(irradiacao.MAR.iloc[0]/1000)
            planilha_e["I20"].value = float(irradiacao.APR.iloc[0]/1000)
            planilha_e["I21"].value = float(irradiacao.MAY.iloc[0]/1000)
            planilha_e["I22"].value = float(irradiacao.JUN.iloc[0]/1000)
            planilha_e["I23"].value = float(irradiacao.JUL.iloc[0]/1000)
            planilha_e["I24"].value = float(irradiacao.AUG.iloc[0]/1000)
            planilha_e["I25"].value = float(irradiacao.SEP.iloc[0]/1000)
            planilha_e["I26"].value = float(irradiacao.OCT.iloc[0]/1000)
            planilha_e["I27"].value = float(irradiacao.NOV.iloc[0]/1000)
            planilha_e["I28"].value = float(irradiacao.DEC.iloc[0]/1000)
            
            # Salva planilha e atualiza front-end
            ex.salva_backend()
            
            # Recarrega planilha de leitura, após ter sido atualizada e recalculada
            wb_l = ex.abre_backend_leitura()
            planilha_l = wb_l["APE"]

            # Finaliza spinner
            next(sp, None)

    # CONSUMOS UC
    dados_iniciais = {
            "Meses": ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
              "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"],
            "HFP": [0]*12,
            "HP": [0]*12
    }
    df_consumos = pd.DataFrame(dados_iniciais)

    # UNIDADE CONSUMIDORA
    with tab2:
        st.write("")

        with st.form("form_UC", enter_to_submit=False, border=False):
            with st.container():
                col1, col2 = st.columns(2)
                with col1:   
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:    
                            st.markdown("##### Nr. UC")
                        with coluna2:
                            nr_uc = st.text_input("Nr. UC", label_visibility="collapsed") 
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:    
                            st.markdown("##### Distribuidora")
                        with coluna2:
                            #st.write(tf.get_distrib())
                            #if distrib_list is None or len(distrib_list) == 0:
                            #    st.error("API não disponível.")
                            #    distrib_list = []  # passa lista vazia para evitar crash
                            distrib = st.selectbox("Distribuidora", [], label_visibility="collapsed")
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:
                            st.markdown("##### Demanda (kW)")
                        with coluna2:
                            dem_contr = st.number_input("Demanda", label_visibility="collapsed", step=0.1)
                with col2:
                    #with st.container():
                        #coluna1, coluna2 = st.columns(2)
                        #with coluna1:    
                            #st.markdown("##### Ano tarifas")
                        #with coluna2:
                            #ano_t = st.text_input("Ano", label_visibility="collapsed", value=0, disabled=True)
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:    
                            st.markdown("##### Subgrupo")
                        with coluna2:
                            subgrupo = st.selectbox("Subgrupo", (subg), label_visibility="collapsed")
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:
                            st.markdown("##### Tipo")
                        with coluna2:
                            mod = st.selectbox("Tipo", (modal), label_visibility="collapsed")                            
                    
            with st.container():
                cl1, cl2 = st.columns(2)
                with cl1:
                    coluna1, coluna2 = st.columns(2)
                    with coluna1:
                        st.write("")    

                        # Salva dados de consumo no dataframe                            
                        tabela_consumos = st.data_editor(
                            df_consumos,
                            hide_index=True, 
                            height=457,
                            use_container_width=True,
                            disabled=["Meses"],
                            column_config={
                                "Meses": st.column_config.TextColumn(),
                                "HFP": st.column_config.NumberColumn("HFP (MWh)"),
                                "HP": st.column_config.NumberColumn("HP (MWh)")
                            }
                        )

                        # Botão para salvar formulário
                        submitted2 = st.form_submit_button("Salvar", icon=":material/check:")  

        tarifas = None

        # Recarrega planilha de leitura
        wb_l = ex.abre_backend_leitura()
        planilha_l = wb_l["APE"]
        
        # Puxa ano de análise
        ano_inicial = planilha_l["C9"].value

        # Executa quando botão Salvar é clicado            
        if submitted2:
            # Cria spinner e inicia
            sp = iter(make_spinner("Salvando"))
            next(sp)

            # Salva variáveis UC na planilha
            planilha_e["C13"].value = nr_uc
            planilha_e["F13"].value = (dem_contr/1000)
            planilha_e["C14"].value = distrib   

            # Puxa ano válido mais recente da API
            ano = tf.get_ano_tarifas(distrib, ano_inicial)

            # Obtém tarifas da API
            tarifas = tf.get_tarifas(distrib, ano, subgrupo, mod)

            # Teste erro subgrupo inexistente
            if "erro" in tarifas:
                st.error(tarifas["erro"])
            else:            
                # Salva tarifas da API na planilha
                planilha_e["C33"].value = float(tarifas["TUSD_D"])
                planilha_e["C36"].value = float(tarifas["TUSD_C_HP"])
                planilha_e["C37"].value = float(tarifas["TUSD_C_HFP"])
                planilha_e["F33"].value = float(tarifas["TUSD_G"])
                planilha_e["F36"].value = float(tarifas["TUSD_APE_HP"])
                planilha_e["F37"].value = float(tarifas["TUSD_APE_HFP"])

            # Remove coluna de Meses
            tabela_consumos = tabela_consumos.drop(["Meses"], axis='columns')

            # Converte dataframe e escreve Consumos na planilha
            rows = dataframe_to_rows(tabela_consumos, index=False, header=False)
            for r_idx, row in enumerate(rows, 17): #startRow = 17
                for c_idx, value in enumerate(row, 4): #startCol = 4 (D)
                    planilha_e.cell(row=r_idx, column=c_idx, value=value)

            # Salva planilha e atualiza front-end
            ex.salva_backend()
            
            # Recarrega planilha de leitura, após ter sido atualizada e recalculada
            wb_l = ex.abre_backend_leitura()
            planilha_l = wb_l["APE"]

            # Finaliza spinner
            next(sp, None)

        with coluna2: 
            st.write("")
            
            # Define range para mostrar
            cell_range = planilha_l['F17':'G28']
            data = []
            for row in cell_range:
                linha = [cell.value for cell in row]
                data.append(linha)  

            # Cria dataframe com range do Excel
            tabela_tot = pd.DataFrame(data)

            # Mostra tabela de totais
            st.dataframe(
                tabela_tot,
                hide_index=True,
                height=457,
                use_container_width=True,
                column_config={
                    1: st.column_config.NumberColumn("Total (MWh)"),
                    2: st.column_config.NumberColumn("MWm")
                }
            )

        with cl2:
            st.write("")

            # Remove coluna de MWm
            tabela_tot = tabela_tot.drop(columns=[1])     
            
            # Insere coluna de Meses
            tabela_tot.insert(0,"Meses",["Jan", "Fev", "Mar", "Abr", "Mai", "Jun","Jul", "Ago", "Set", "Out", "Nov", "Dez"])

            # Define headers das colunas
            tabela_tot.columns = ['Meses', 'Total (MWh)']
            
            # Define coluna Meses como índice
            tabela_tot.set_index('Meses', inplace=True)

            plt.figure(figsize=(10, 4.8))
            #sns.barplot(x=tabela_tot.index, y=tabela_tot['Total (MWh)'], palette="viridis")
            plt.bar(tabela_tot.index, tabela_tot['Total (MWh)'], color='#c95f0d')
            plt.gca().set_facecolor('#1a1a1a')
            plt.gca().spines['bottom'].set_color('#cecece')
            plt.gca().spines['left'].set_color('#cecece')
            plt.gca().spines['top'].set_color('#1a1a1a')
            plt.gca().spines['right'].set_color('#1a1a1a')
            plt.gca().grid(True, which='major', axis='y', color='#cecece', linestyle='-', linewidth=0.1)
            plt.gcf().set_facecolor('#1a1a1a')
            plt.tick_params(colors='#cecece')
            plt.title('Consumo Total', color="#cecece")

            # Insere valor em cima de cada coluna
            #for i, value in enumerate(tabela_tot['Total (MWh)']):
            #    plt.text(i, value + 10, f'{value:.2f}', ha='center', fontsize=8, color='white')  # Texto branco

            st.pyplot(plt)            
            #st.bar_chart(tabela_tot, x_label="Meses") 

    # USINA FV
    with tab3:
        st.write("")

        with st.form("form_UFV", enter_to_submit=False, border=False):
            with st.container():
                col1, col2 = st.columns(2)
                with col1:   
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:    
                            st.markdown("##### Pot. Inversor (kW)")
                        with coluna2:
                            pinv = st.number_input("Pot. Inversor", label_visibility="collapsed") 
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:    
                            st.markdown("##### Tipo Instalação")
                        with coluna2:
                            tipo_inst = st.selectbox("Tipo Instalação", (estr), label_visibility="collapsed")
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:
                            st.markdown("##### FP")
                        with coluna2:
                            fp = st.selectbox("FP", (fatorp), label_visibility="collapsed")                    
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna2:
                            # Botão para salvar formulário
                            submitted3 = st.form_submit_button("Salvar", icon=":material/check:") 

            with st.container():
                cl1, cl2 = st.columns(2)
                with cl1:
                    coluna1, coluna2 = st.columns(2)   

        # Executa quando botão Salvar é clicado 
        if submitted3:       
            # Cria spinner e inicia
            sp = iter(make_spinner("Salvando"))
            next(sp)

            # Salva variáveis na planilha
            planilha_e["J14"].value = (pinv/1000)
            planilha_e["M13"].value = tipo_inst
            planilha_e["M14"].value = fp

            # Salva planilha e atualiza front-end
            ex.salva_backend()
            
            # Recarrega planilha de leitura, após ter sido atualizada e recalculada
            wb_l = ex.abre_backend_leitura()
            planilha_l = wb_l["APE"]

            # Finaliza spinner
            next(sp, None)

        with col2:
            #st.write("")
            with st.container():
                c1, c2 = st.columns(2)
                with c2:    
                    st.metric(label="Pot Pico calculada", value=millify((planilha_l["J13"].value)*1000000, precision=2, prefixes=[' kWp',' MWp']))
                    st.metric(label="Pot Pico projeto", value=millify((planilha_l["K14"].value)*1000000, precision=2, prefixes=[' kWp',' MWp']))

        with coluna2: 
            st.write("")
            
            # Define range para mostrar
            cell_range2 = planilha_l['J17':'K28']
            data2 = []
            for row in cell_range2:
                linha = [cell.value for cell in row]
                data2.append(linha)  

            # Cria dataframe com range do Excel
            tabela_ger = pd.DataFrame(data2)

            # Mostra tabela de totais
            st.dataframe(
                tabela_ger,
                hide_index=True,
                height=457,
                use_container_width=True,
                column_config={
                    1: st.column_config.NumberColumn("Geração (MWh)"),
                    2: st.column_config.NumberColumn("MWm")
                }
            )

        with cl2:
            st.write("")

            # Remove coluna de MWm
            tabela_ger = tabela_ger.drop(columns=[1])     
            
            # Insere coluna de Meses
            tabela_ger.insert(0,"Meses",["Jan", "Fev", "Mar", "Abr", "Mai", "Jun","Jul", "Ago", "Set", "Out", "Nov", "Dez"])

            # Define headers das colunas
            tabela_ger.columns = ['Meses', 'Geração (MWh)']
            
            # Define coluna Meses como índice
            tabela_ger.set_index('Meses', inplace=True)

            plt.figure(figsize=(10, 4.8))
            #sns.barplot(x=tabela_tot.index, y=tabela_tot['Total (MWh)'], palette="viridis")
            plt.bar(tabela_ger.index, tabela_ger['Geração (MWh)'], color='#4da13a')
            plt.gca().set_facecolor('#1a1a1a')
            plt.gca().spines['bottom'].set_color('#cecece')
            plt.gca().spines['left'].set_color('#cecece')
            plt.gca().spines['top'].set_color('#1a1a1a')
            plt.gca().spines['right'].set_color('#1a1a1a')
            plt.gca().grid(True, which='major', axis='y', color='#cecece', linestyle='-', linewidth=0.1)
            plt.gcf().set_facecolor('#1a1a1a')
            plt.tick_params(colors='#cecece')
            plt.title('Geração Total', color="#cecece")

            # Insere valor em cima de cada coluna
            #for i, value in enumerate(tabela_tot['Total (MWh)']):
            #    plt.text(i, value + 10, f'{value:.2f}', ha='center', fontsize=8, color='white')  # Texto branco

            st.pyplot(plt)            
            #st.bar_chart(tabela_tot, x_label="Meses") 

    # CONTRATO ENERGIA
    dados_iniciais2 = {
            "Anos": [ano_analise, ano_analise+1, ano_analise+2, ano_analise+3, ano_analise+4, ano_analise+5, ano_analise+6, ano_analise+7,
              ano_analise+8, ano_analise+9, ano_analise+10, ano_analise+11, ano_analise+12, ano_analise+13, ano_analise+14],
            "Preço contrato": [0]*15,
            "PLD": [0]*15
    }
    df_contrato = pd.DataFrame(dados_iniciais2)

    # CONTRATO DE ENERGIA
    with tab4:
        st.write("")

        with st.form("form_contratoEnergia", enter_to_submit=False, border=False):
            with st.container():
                col1, col2 = st.columns(2)
                with col1:   
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:    
                            st.markdown("##### Volume (MWm)")
                        with coluna2:
                            volume = st.number_input("Volume", label_visibility="collapsed")
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:    
                            st.markdown("##### Sazonalização (%)")
                        with coluna2:
                            sazonalizacao = st.number_input("Sazonalização", label_visibility="collapsed", step=1)
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:    
                            st.markdown("##### Flex máx (%)")
                        with coluna2:
                            flex_max = st.number_input("Flex máx", label_visibility="collapsed", value=0)
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:    
                            st.markdown("##### Flex mín (%)")
                        with coluna2:
                            flex_min = st.number_input("Flex mín", label_visibility="collapsed", value=0)
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:    
                            st.markdown("##### Prazo (ano)")
                        with coluna2:
                            prazo_contr = st.number_input("Prazo", label_visibility="collapsed", value=0)
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:    
                            st.markdown("##### Desconto (%)")
                        with coluna2:
                            desconto = st.number_input("Desconto", label_visibility="collapsed", step=1, value=50)
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna2:                            
                            # Botão para salvar formulário
                            submitted4 = st.form_submit_button("Salvar", icon=":material/check:")                              
                with col2:
                    # Salva dados de preços no dataframe                            
                    tabela_contrato = st.data_editor(
                        df_contrato,
                        hide_index=True, 
                        height=380,
                        width=450,
                        disabled=["Anos"],
                        column_config={
                            "Anos": st.column_config.TextColumn(),
                            "Preço contrato": st.column_config.NumberColumn("Preço Contrato"),
                            "PLD": st.column_config.NumberColumn("PLD (R$/MWh)")
                        }
                    )   

        # Executa quando botão Salvar é clicado            
        if submitted4:
            # Cria spinner e inicia
            sp = iter(make_spinner("Salvando"))
            next(sp)
            
            # Salva variáveis na planilha
            planilha_e["J33"].value = volume
            planilha_e["J34"].value = (sazonalizacao/100)
            planilha_e["J35"].value = (flex_max/100)
            planilha_e["J36"].value = (flex_min/100)
            planilha_e["J37"].value = prazo_contr
            planilha_e["J38"].value = (desconto/100)

            # Remove coluna de Anos
            tabela_contrato = tabela_contrato.drop(["Anos"], axis='columns')

            # Converte dataframe e escreve Preços na planilha
            rows2 = dataframe_to_rows(tabela_contrato, index=False, header=False)
            for r_idx, row in enumerate(rows2, 34): #startRow = 34
                for c_idx, value in enumerate(row, 12): #startCol = 12 (L)
                    planilha_e.cell(row=r_idx, column=c_idx, value=value)

            # Salva planilha e atualiza front-end
            ex.salva_backend()

            # Recarrega planilha de leitura, após ter sido atualizada e recalculada
            wb_l = ex.abre_backend_leitura()
            planilha_l = wb_l["APE"]

            # Finaliza spinner
            next(sp, None)

        with st.container():
            cl1, cl2 = st.columns(2)
            with cl1:
                st.write("")
            
                # Define range para mostrar
                cell_range3 = planilha_l['B44':'G55']
                data3 = []
                for row in cell_range3:
                    linha = [cell.value for cell in row]
                    data3.append(linha)  

                # Cria dataframe com range do Excel
                tabela_contr_saz = pd.DataFrame(data3)

                # Mostra tabela de contrato sazonalizado
                st.dataframe(
                    tabela_contr_saz,
                    hide_index=True,
                    height=457,
                    width=900,
                    column_config={
                        1: st.column_config.NumberColumn("Consumo (MWm)"),
                        2: st.column_config.NumberColumn("Contrato Flat (MWm)"),
                        3: st.column_config.NumberColumn("Contrato Sazo (MWm)"),
                        4: st.column_config.NumberColumn("Flex máx (MWm)"),
                        5: st.column_config.NumberColumn("Flex mín (MWm)"),
                        6: st.column_config.NumberColumn("Consumo c/ UFV (MWm)")
                    }
                )
            with cl2:
                st.write("")
                
                # Insere coluna de Meses
                tabela_contr_saz.insert(0,"Meses",["Jan", "Fev", "Mar", "Abr", "Mai", "Jun","Jul", "Ago", "Set", "Out", "Nov", "Dez"])

                # Define headers das colunas
                tabela_contr_saz.columns = ['Meses', 'Consumo (MWm)', 'Contrato Flat (MWm)', 'Contrato Sazo (MWm)', 'Flex máx (MWm)', 'Flex mín (MWm)', 'Consumo c/ UFV (MWm)']
                
                # Define coluna Meses como índice
                tabela_contr_saz.set_index('Meses', inplace=True)

                plt.figure(figsize=(10, 3.65))
                plt.plot(tabela_contr_saz.index, tabela_contr_saz['Consumo (MWm)'], color='#c95f0d', label='Consumo')
                plt.plot(tabela_contr_saz.index, tabela_contr_saz['Contrato Flat (MWm)'], color='white', label='Contrato Flat')
                plt.plot(tabela_contr_saz.index, tabela_contr_saz['Contrato Sazo (MWm)'], color='#5c7190', label='Contrato Sazo')
                plt.plot(tabela_contr_saz.index, tabela_contr_saz['Consumo c/ UFV (MWm)'], color='green', label='Consumo c/ UFV')
                plt.fill_between(tabela_contr_saz.index, tabela_contr_saz['Flex máx (MWm)'], tabela_contr_saz['Flex mín (MWm)'], color='#afb3bb', label='Flexibilidade')

                plt.gca().set_facecolor('#1a1a1a')
                plt.gca().spines['bottom'].set_color('#cecece')
                plt.gca().spines['left'].set_color('#cecece')
                plt.gca().spines['top'].set_color('#1a1a1a')
                plt.gca().spines['right'].set_color('#1a1a1a')
                plt.gca().grid(True, which='major', axis='y', color='#cecece', linestyle='-', linewidth=0.1)
                plt.gcf().set_facecolor('#1a1a1a')
                plt.tick_params(colors='#cecece')
                plt.title('Contrato Sazonalizado + UFV', color="#cecece")

                plt.legend(loc="upper center", bbox_to_anchor=(0.5, -0.15), ncol=2, facecolor='#1a1a1a', edgecolor='#1a1a1a', labelcolor='#cecece')

                st.pyplot(plt)            

    # VIABILIDADE
    with tab5:
        st.write("")

        with st.form("form_viabilidade", enter_to_submit=False, border=False):
            with st.container():
                col1, col2 = st.columns(2)
                with col1:   
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:    
                            st.markdown("##### Investimento")
                        with coluna2:
                            # Mostra valor do sistema calculado na planilha
                            investimento = st.text_input("Investimento", label_visibility="collapsed", value=f'R$ {(planilha_l["J5"].value):,.2f}'.replace(",", "X").replace(".", ",").replace("X", "."), disabled=True)
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:    
                            st.markdown("##### Inflação (a.a.)")
                        with coluna2:
                            inflacao = st.number_input("Inflação", label_visibility="collapsed", step=0.1, value=10.0)
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:    
                            st.markdown("##### O&M (a.a.)")
                        with coluna2:
                            oem = st.number_input("O&M", label_visibility="collapsed", step=0.1, value=0.5)
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:    
                            st.markdown("##### TMA (a.a.)")
                        with coluna2:
                            tma = st.number_input("TMA", label_visibility="collapsed", step=0.1, value=5.0)
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:    
                            st.markdown("##### Tipo Pgto")
                        with coluna2:
                            tipo_pgto = st.selectbox("Tipo Pgto", (tipopag), label_visibility="collapsed")
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:    
                            st.markdown("##### Percentual Financiado")
                        with coluna2:
                            perc_financ = st.number_input("Perc financiado", label_visibility="collapsed", step=0.1)
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:    
                            st.markdown("##### Tipo Financiamento")
                        with coluna2:
                            tipo_financ = st.selectbox("Tipo Financ", (tipofinanc), label_visibility="collapsed")
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:    
                            st.markdown("##### Prazo (anos)")
                        with coluna2:
                            prazo_financ = st.number_input("Prazo", label_visibility="collapsed", value=0)
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:    
                            st.markdown("##### Carência (anos)")
                        with coluna2:
                            car_financ = st.number_input("Carência", label_visibility="collapsed", value=0)
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna1:    
                            st.markdown("##### Taxa (a.a.)")
                        with coluna2:
                            taxa_financ = st.number_input("Taxa", label_visibility="collapsed", step=0.01)
                    with st.container():
                        coluna1, coluna2 = st.columns(2)
                        with coluna2:
                            # Botão para salvar formulário
                            submitted5 = st.form_submit_button("Calcular", type="primary", icon=":material/bolt:") 

        # Executa quando botão Calcular é clicado 
        if submitted5:       
            # Cria spinner e inicia
            sp = iter(make_spinner("Salvando"))
            next(sp)

            # Salva variáveis na planilha
            #planilha_e["J5"].value = investimento
            planilha_e["J6"].value = (inflacao/100)
            planilha_e["J7"].value = (oem/100)
            planilha_e["J8"].value = (tma/100)
            planilha_e["J9"].value = tipo_pgto
            planilha_e["N5"].value = (perc_financ/100)
            planilha_e["N6"].value = tipo_financ
            planilha_e["N7"].value = prazo_financ
            planilha_e["N8"].value = car_financ
            planilha_e["N9"].value = (taxa_financ/100)

            # Salva planilha e atualiza front-end
            ex.salva_backend()
            
            # Recarrega planilha de leitura, após ter sido atualizada e recalculada
            wb_l = ex.abre_backend_leitura()
            planilha_l = wb_l["APE"]

            # Finaliza spinner
            next(sp, None)

        with col2:
            with st.container():
                coluna1, coluna2 = st.columns(2)
                with coluna1:    
                    st.metric(label="Valor Financiado", value=f'R$ {(planilha_l["Q5"].value):,.2f}'.replace(",", "X").replace(".", ",").replace("X", ".")) 
                with coluna2:
                    st.metric(label="Entrada", value=f'R$ {(planilha_l["Q6"].value):,.2f}'.replace(",", "X").replace(".", ",").replace("X", ".")) 
            st.write("")
            st.write("")
            st.write("")
            st.write("")
            st.write("")
            st.divider()
            with st.container():
                coluna1, coluna2 = st.columns(2)
                with coluna1:
                    st.metric(label="TIR (IPCA+)", value=f'{(planilha_l["Q13"].value):.2%}'.replace(".", ","))
                with coluna2:
                    st.metric(label="VPL", value=f'R$ {(planilha_l["Q14"].value):,.2f}'.replace(",", "X").replace(".", ",").replace("X", "."))                   
            with st.container():
                coluna1, coluna2 = st.columns(2)
                with coluna1:    
                    st.metric(label="LCOE", value=planilha_l["Q16"].value)
                with coluna2:
                    st.metric(label="PAYBACK", value=planilha_l["Q17"].value)                    
            with st.container():
                coluna1, coluna2 = st.columns(2)
                with coluna1:    
                    st.metric(label="CAIXA TOTAL", value=f'R$ {(planilha_l["Q15"].value):,.2f}'.replace(",", "X").replace(".", ",").replace("X", "."))  
            st.divider()

        with st.container():
            cl1, cl2 = st.columns(2)
            with cl1:
                st.write("")
            
                # Define range para mostrar
                cell_range4 = planilha_l['O20':'S50']
                data4 = []
                for row in cell_range4:
                    linha = [cell.value for cell in row]
                    data4.append(linha)  

                # Cria dataframe com range do Excel
                tabela_res = pd.DataFrame(data4)

                # Substitui valores NaN por 0
                tabela_res = tabela_res.fillna(0)

                # Converte coluna de Anos para string
                tabela_res[0] = tabela_res[0].astype(str)

                # Mostra tabela de resultado econômico
                st.dataframe(
                    tabela_res,
                    hide_index=True,
                    height=590,
                    use_container_width=True,
                    column_config={
                        1: st.column_config.NumberColumn("Anos", format="%d"),
                        2: st.column_config.NumberColumn("Economia", format="R$ %.2f"),
                        3: st.column_config.NumberColumn("Parcelas Pgto", format="R$ %.2f"),
                        4: st.column_config.NumberColumn("Fluxo de Caixa", format="R$ %.2f"),
                        5: st.column_config.NumberColumn("FC Acumulado", format="R$ %.2f")
                    }
                )

            with cl2:
                st.write("")

                # Remove colunas desnecessárias ao gráfico
                tabela_res = tabela_res.drop(columns=[1, 2, 3])  

                # Define headers das colunas
                tabela_res.columns = ['Anos', 'FC Acumulado']
                
                # Define coluna Anos como índice
                tabela_res.set_index('Anos', inplace=True)

                plt.figure(figsize=(10, 6))
                plt.bar(tabela_res.index, tabela_res['FC Acumulado'], color='#3c7bf8')
                plt.gca().set_facecolor('#1a1a1a')
                plt.gca().spines['bottom'].set_color('#cecece')
                plt.gca().spines['left'].set_color('#cecece')
                plt.gca().spines['top'].set_color('#1a1a1a')
                plt.gca().spines['right'].set_color('#1a1a1a')
                plt.gca().grid(True, which='major', axis='y', color='#cecece', linestyle='-', linewidth=0.1)
                plt.gcf().set_facecolor('#1a1a1a')
                plt.tick_params(colors='#cecece')
                plt.xticks(rotation=90)
                plt.title('Fluxo de Caixa Acumulado', color="#cecece")

                st.pyplot(plt)  

    # CONSUMOS ZERO
    consumos_zero = {
            "HFP": [0]*12,
            "HP": [0]*12
    }
    df_cons_zero = pd.DataFrame(consumos_zero)

    # CONTRATO ZERO
    contrato_zero = {
            "Preço contrato": [0]*15,
            "PLD": [0]*15
    }
    df_contr_zero = pd.DataFrame(contrato_zero)

    # RESUMO             
    with tab6:
        st.write("--- Gerar relatório em PDF ---")

        reiniciar = st.button("Reiniciar")
        if reiniciar:
            
            # Cria spinner e inicia
            sp = iter(make_spinner("Finalizando"))
            next(sp)

            # Limpa campos da planilha
            planilha_e["C5"].value = ""
            planilha_e["C6"].value = ""
            planilha_e["C7"].value = ""
            planilha_e["C9"].value = ""
            planilha_e["C13"].value = ""
            planilha_e["C14"].value = ""
            planilha_e["F13"].value = ""
            
            # Converte dataframe e zera Consumos na planilha
            rows = dataframe_to_rows(df_cons_zero, index=False, header=False)
            for r_idx, row in enumerate(rows, 17): #startRow = 17
                for c_idx, value in enumerate(row, 4): #startCol = 4 (D)
                    planilha_e.cell(row=r_idx, column=c_idx, value=value)

            planilha_e["C33"].value = ""
            planilha_e["C36"].value = ""
            planilha_e["C37"].value = ""
            planilha_e["F33"].value = ""
            planilha_e["F36"].value = ""
            planilha_e["F37"].value = ""

            planilha_e["J6"].value = ""
            planilha_e["J7"].value = ""
            planilha_e["J8"].value = ""
            planilha_e["N5"].value = ""
            planilha_e["N7"].value = ""
            planilha_e["N8"].value = ""
            planilha_e["N9"].value = ""

            planilha_e["J14"].value = ""

            planilha_e["J33"].value = ""
            planilha_e["J34"].value = ""
            planilha_e["J35"].value = ""
            planilha_e["J36"].value = ""
            planilha_e["J37"].value = ""
            planilha_e["J38"].value = ""

            # Converte dataframe e zera Preços na planilha
            rows2 = dataframe_to_rows(df_contr_zero, index=False, header=False)
            for r_idx, row in enumerate(rows2, 34): #startRow = 34
                for c_idx, value in enumerate(row, 12): #startCol = 12 (L)
                    planilha_e.cell(row=r_idx, column=c_idx, value=value)

            # Salva planilha e atualiza front-end
            ex.salva_backend()
            
            # Recarrega planilha de leitura, após ter sido atualizada e recalculada
            wb_l = ex.abre_backend_leitura()
            planilha_l = wb_l["APE"]

            # Limpa dados no front-end
            #tabela_tot = None
            #tabela_ger = None
            #tabela_contr_saz = None
            #tabela_res = None

            keyboard.press_and_release('f5')

            # Finaliza spinner
            next(sp, None)

# RODAPÉ DO APP
with st.container():
    st.write("")
    st.divider()
    with st.container():
        col1,col2,col3 = st.columns([25,10,20])
        with col2:            
            st.markdown(":grey[v1.0 (2024)  |  by CS]")
                
            
                

    